from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet1", 0)
ws2 = wb.create_sheet("Mysheet2", 1)
ws3 = wb.create_sheet("Mysheet3", 2)
wss1 = wb['Mysheet1']
wss2 = wb['Mysheet2']
wss3 = wb['Mysheet3']
n=wb.sheetnames
wb.remove(wb[n[3]])

wss1.column_dimensions['A'].width = 15
wss1.column_dimensions['B'].width = 15
wss1.column_dimensions['C'].width = 15
wss1.column_dimensions['D'].width = 15

wss1['A1'] = 'Type'
wss1['B1'] = 'Price'
wss1['C1'] = 'SalePrice'
wss1['D1'] = 'Savings'

currentCell1 = wss1['A1']
currentCell1.alignment = Alignment(horizontal='center')
currentCell2 = wss1['B1']
currentCell2.alignment = Alignment(horizontal='center')
currentCell3 = wss1['C1']
currentCell3.alignment = Alignment(horizontal='center')
currentCell4 = wss1['D1']
currentCell4.alignment = Alignment(horizontal='center')

wss1['A2'] = 'Apples'
wss1['A3'] = 'Oranges'
wss1['A4'] = 'Bananas'

wb.save('example2.xlsx')




